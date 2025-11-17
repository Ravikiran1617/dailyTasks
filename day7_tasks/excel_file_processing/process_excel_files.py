import pandas as pd
from openpyxl import load_workbook

# 1️⃣ Load the original Excel file with multiple sheets
excel_file = "Book1.xlsx"
all_sheets = pd.read_excel(excel_file, sheet_name=None)  # Load all sheets

processed_sheets = {}

for sheet_name, df in all_sheets.items():
    print(f"Processing sheet: {sheet_name}")
    
    # 2️⃣ Clean data
    df = df.drop_duplicates()  # Remove duplicate rows
    
    # Fill missing numeric columns with 0 (or median if preferred)
    for col in df.select_dtypes(include='number'):
        df[col] = df[col].fillna(0)
    
    # Fill missing text columns with empty string
    for col in df.select_dtypes(include='object'):
        df[col] = df[col].fillna("")
    
    # Trim spaces & normalize text
    for col in df.select_dtypes(include='object'):
        df[col] = df[col].str.strip().str.title()
    
    # Safe numeric conversion for text columns
    for col in df.select_dtypes(include='object'):
        df[col] = pd.to_numeric(df[col], errors='ignore')  # ignore if cannot convert
    
    processed_sheets[sheet_name] = df

# 3️⃣ Save cleaned sheets to a new Excel
cleaned_file = "cleaned_data.xlsx"
with pd.ExcelWriter(cleaned_file) as writer:
    for sheet_name, df in processed_sheets.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print("✅ Cleaning complete. Saved to 'cleaned_data.xlsx'.")

# 4️⃣ Resize columns and rows for neat output
wb = load_workbook(cleaned_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2  # add padding
    
    # Auto-adjust row heights for multi-line text
    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                max_lines = max(max_lines, cell.value.count('\n') + 1)
        ws.row_dimensions[row[0].row].height = max_lines * 15

# Save final nicely formatted Excel
final_file = "cleaned_data_resized.xlsx"
wb.save(final_file)
print(f"✅ All sheets resized and saved to '{final_file}'.")  
