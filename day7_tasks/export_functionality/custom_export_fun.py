import pandas as pd
import numpy as np
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================================
# SAMPLE DATA FOR DEMONSTRATION
# ============================================================================
# Create sample customer data
data = {
    'Customer_ID': ['C001', 'C002', 'C003', 'C004', 'C005'],
    'Name': ['John Smith', 'Jane Doe', 'Bob Johnson', 'Alice Williams', 'Charlie Brown'],
    'Age': [28, 34, 45, 29, 52],
    'City': ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix'],
    'Salary': [55000, 68000, 72000, 51000, 89000],
    'Department': ['Sales', 'IT', 'HR', 'Sales', 'IT'],
    'Join_Date': pd.date_range('2020-01-01', periods=5, freq='M')
}
df = pd.DataFrame(data)

print("="*70)
print("DATA EXPORT FUNCTIONALITIES - COMPLETE GUIDE")
print("="*70)
print("\nSample Dataset:")
print(df)
print("\n")

# Create output directory
os.makedirs('exports', exist_ok=True)

# ============================================================================
# PART 1: CSV EXPORT
# ============================================================================
print("="*70)
print("PART 1: CSV EXPORT")
print("="*70)

# METHOD 1: Basic CSV Export
print("\n1️⃣ Basic CSV Export")
print("-" * 70)
csv_basic = 'exports/data_basic.csv'
df.to_csv(csv_basic, index=False)
print(f"✓ Exported to: {csv_basic}")
print(f"  - No index column")
print(f"  - Default separator: comma (,)")

# METHOD 2: CSV with Custom Separator
print("\n2️⃣ CSV with Custom Separator")
print("-" * 70)
csv_custom_sep = 'exports/data_semicolon.csv'
df.to_csv(csv_custom_sep, index=False, sep=';')
print(f"✓ Exported to: {csv_custom_sep}")
print(f"  - Separator: semicolon (;)")
print(f"  - Useful for European Excel formats")

# METHOD 3: CSV with Selected Columns
print("\n3️⃣ CSV with Selected Columns Only")
print("-" * 70)
csv_selected = 'exports/data_selected_columns.csv'
df[['Customer_ID', 'Name', 'Salary']].to_csv(csv_selected, index=False)
print(f"✓ Exported to: {csv_selected}")
print(f"  - Only columns: Customer_ID, Name, Salary")

# METHOD 4: CSV with Custom Encoding
print("\n4️⃣ CSV with UTF-8 BOM Encoding")
print("-" * 70)
csv_utf8_bom = 'exports/data_utf8_bom.csv'
df.to_csv(csv_utf8_bom, index=False, encoding='utf-8-sig')
print(f"✓ Exported to: {csv_utf8_bom}")
print(f"  - Encoding: UTF-8 with BOM")
print(f"  - Better Excel compatibility for special characters")

# METHOD 5: CSV with Filtered Data
print("\n5️⃣ CSV with Filtered Data")
print("-" * 70)
csv_filtered = 'exports/data_high_salary.csv'
df[df['Salary'] > 60000].to_csv(csv_filtered, index=False)
print(f"✓ Exported to: {csv_filtered}")
print(f"  - Only rows where Salary > 60000")

# METHOD 6: CSV with Custom Date Format
print("\n6️⃣ CSV with Custom Date Format")
print("-" * 70)
csv_date_format = 'exports/data_custom_dates.csv'
df.to_csv(csv_date_format, index=False, date_format='%Y-%m-%d')
print(f"✓ Exported to: {csv_date_format}")
print(f"  - Date format: YYYY-MM-DD")

# METHOD 7: CSV Append Mode
print("\n7️⃣ CSV Append Mode (Add to Existing File)")
print("-" * 70)
csv_append = 'exports/data_append.csv'
# First write
df.head(2).to_csv(csv_append, index=False, mode='w')
# Then append
df.tail(2).to_csv(csv_append, index=False, mode='a', header=False)
print(f"✓ Exported to: {csv_append}")
print(f"  - First 2 rows written, then last 2 appended")

# ============================================================================
# PART 2: EXCEL EXPORT
# ============================================================================
print("\n" + "="*70)
print("PART 2: EXCEL EXPORT")
print("="*70)

# METHOD 1: Basic Excel Export
print("\n1️⃣ Basic Excel Export")
print("-" * 70)
excel_basic = 'exports/data_basic.xlsx'
df.to_excel(excel_basic, index=False, sheet_name='Customers')
print(f"✓ Exported to: {excel_basic}")
print(f"  - Sheet name: Customers")
print(f"  - No index column")

# METHOD 2: Excel with Multiple Sheets
print("\n2️⃣ Excel with Multiple Sheets")
print("-" * 70)
excel_multi = 'exports/data_multiple_sheets.xlsx'
with pd.ExcelWriter(excel_multi, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='All_Data', index=False)
    df[df['Department'] == 'Sales'].to_excel(writer, sheet_name='Sales', index=False)
    df[df['Department'] == 'IT'].to_excel(writer, sheet_name='IT', index=False)
    
    # Add a summary sheet
    summary = df.groupby('Department').agg({
        'Salary': ['mean', 'min', 'max'],
        'Customer_ID': 'count'
    }).round(2)
    summary.to_excel(writer, sheet_name='Summary')

print(f"✓ Exported to: {excel_multi}")
print(f"  - Sheet 1: All_Data (all records)")
print(f"  - Sheet 2: Sales (filtered)")
print(f"  - Sheet 3: IT (filtered)")
print(f"  - Sheet 4: Summary (aggregated)")

# METHOD 3: Excel with Formatting
print("\n3️⃣ Excel with Advanced Formatting")
print("-" * 70)
excel_formatted = 'exports/data_formatted.xlsx'

# Create Excel file with formatting
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Customers"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center")

# Border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers with formatting
for col_num, column_title in enumerate(df.columns, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = column_title
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
for row_num, row_data in enumerate(df.values, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.border = thin_border
        
        # Align numbers to right
        if isinstance(cell_value, (int, float)):
            cell.alignment = Alignment(horizontal="right")

# Auto-adjust column widths
for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column_letter].width = adjusted_width

workbook.save(excel_formatted)
print(f"✓ Exported to: {excel_formatted}")
print(f"  - Blue header with white text")
print(f"  - Borders on all cells")
print(f"  - Auto-adjusted column widths")
print(f"  - Numbers aligned right")

# METHOD 4: Excel with Conditional Formatting
print("\n4️⃣ Excel with Conditional Formatting (Color Coding)")
print("-" * 70)
excel_conditional = 'exports/data_conditional.xlsx'

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Color Coded Data"

# Write data
for r in dataframe_to_rows(df, index=False, header=True):
    sheet.append(r)

# Apply conditional formatting to Salary column (column E, index 5)
from openpyxl.styles import Color
for row in range(2, sheet.max_row + 1):
    salary_cell = sheet.cell(row=row, column=5)
    salary = salary_cell.value
    
    # Color code based on salary
    if salary < 60000:
        salary_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    elif salary < 75000:
        salary_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    else:
        salary_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green

workbook.save(excel_conditional)
print(f"✓ Exported to: {excel_conditional}")
print(f"  - Red: Salary < $60,000")
print(f"  - Yellow: Salary $60,000-$75,000")
print(f"  - Green: Salary > $75,000")

# METHOD 5: Excel with Formulas
print("\n5️⃣ Excel with Formulas")
print("-" * 70)
excel_formulas = 'exports/data_with_formulas.xlsx'

workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers
headers = list(df.columns) + ['Tax (30%)', 'Net Salary']
sheet.append(headers)

# Write data with formulas
for idx, row in df.iterrows():
    row_data = list(row)
    sheet.append(row_data)
    
    current_row = sheet.max_row
    # Add formula for Tax (30% of Salary)
    sheet.cell(row=current_row, column=len(df.columns) + 1).value = f"=E{current_row}*0.3"
    # Add formula for Net Salary (Salary - Tax)
    sheet.cell(row=current_row, column=len(df.columns) + 2).value = f"=E{current_row}-G{current_row}"

workbook.save(excel_formulas)
print(f"✓ Exported to: {excel_formulas}")
print(f"  - Added 'Tax (30%)' column with formula")
print(f"  - Added 'Net Salary' column with formula")

# ============================================================================
# PART 3: JSON EXPORT
# ============================================================================
print("\n" + "="*70)
print("PART 3: JSON EXPORT")
print("="*70)

# METHOD 1: Basic JSON Export (Records Format)
print("\n1️⃣ Basic JSON Export - Records Format")
print("-" * 70)
json_records = 'exports/data_records.json'
df.to_json(json_records, orient='records', indent=2, date_format='iso')
print(f"✓ Exported to: {json_records}")
print(f"  - Format: Array of objects")
print(f"  - Structure: [{{'col1': 'val1'}}, {{'col1': 'val2'}}]")
print("\nPreview:")
print(df.head(2).to_json(orient='records', indent=2, date_format='iso'))

# METHOD 2: JSON with Index
print("\n2️⃣ JSON Export - Index Format")
print("-" * 70)
json_index = 'exports/data_index.json'
df.to_json(json_index, orient='index', indent=2, date_format='iso')
print(f"✓ Exported to: {json_index}")
print(f"  - Format: Object with index keys")
print(f"  - Structure: {{'0': {{'col1': 'val1'}}, '1': {{'col1': 'val2'}}}}")

# METHOD 3: JSON Columns Format
print("\n3️⃣ JSON Export - Columns Format")
print("-" * 70)
json_columns = 'exports/data_columns.json'
df.to_json(json_columns, orient='columns', indent=2, date_format='iso')
print(f"✓ Exported to: {json_columns}")
print(f"  - Format: Object with column keys")
print(f"  - Structure: {{'col1': {{'0': 'val1', '1': 'val2'}}}}")

# METHOD 4: JSON with Custom Structure
print("\n4️⃣ JSON Export - Custom Structure")
print("-" * 70)
json_custom = 'exports/data_custom_structure.json'

custom_data = {
    'metadata': {
        'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_records': len(df),
        'columns': list(df.columns)
    },
    'data': df.to_dict(orient='records'),
    'summary': {
        'avg_salary': float(df['Salary'].mean()),
        'total_employees': len(df),
        'departments': df['Department'].unique().tolist()
    }
}

with open(json_custom, 'w') as f:
    json.dump(custom_data, f, indent=2, default=str)

print(f"✓ Exported to: {json_custom}")
print(f"  - Includes metadata")
print(f"  - Includes summary statistics")
print(f"  - Custom nested structure")

# METHOD 5: Nested JSON (Grouped Data)
print("\n5️⃣ JSON Export - Nested/Grouped Format")
print("-" * 70)
json_nested = 'exports/data_nested.json'

nested_data = {}
for dept in df['Department'].unique():
    dept_data = df[df['Department'] == dept].to_dict(orient='records')
    nested_data[dept] = {
        'count': len(dept_data),
        'avg_salary': float(df[df['Department'] == dept]['Salary'].mean()),
        'employees': dept_data
    }

with open(json_nested, 'w') as f:
    json.dump(nested_data, f, indent=2, default=str)

print(f"✓ Exported to: {json_nested}")
print(f"  - Grouped by Department")
print(f"  - Nested structure with statistics")

# METHOD 6: JSON Lines Format (JSONL)
print("\n6️⃣ JSON Lines Export (JSONL) - For Big Data")
print("-" * 70)
json_lines = 'exports/data.jsonl'
df.to_json(json_lines, orient='records', lines=True, date_format='iso')
print(f"✓ Exported to: {json_lines}")
print(f"  - Format: One JSON object per line")
print(f"  - Ideal for streaming/big data")
print(f"  - Used by: MongoDB, Elasticsearch")

# ============================================================================
# PART 4: ADVANCED DATA EXPORTER CLASS
# ============================================================================
print("\n" + "="*70)
print("PART 4: REUSABLE DATA EXPORTER CLASS")
print("="*70)

class DataExporter:
    """
    Comprehensive data exporter with multiple format support
    """
    
    def __init__(self, dataframe, output_dir='exports'):
        self.df = dataframe
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export(self, filename, format='csv', **kwargs):
        """
        Universal export method
        
        Args:
            filename: Output filename (without extension)
            format: 'csv', 'excel', 'json', or 'all'
            **kwargs: Format-specific options
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'csv' or format == 'all':
            filepath = os.path.join(self.output_dir, f"{filename}.csv")
            self.export_csv(filepath, **kwargs)
        
        if format == 'excel' or format == 'all':
            filepath = os.path.join(self.output_dir, f"{filename}.xlsx")
            self.export_excel(filepath, **kwargs)
        
        if format == 'json' or format == 'all':
            filepath = os.path.join(self.output_dir, f"{filename}.json")
            self.export_json(filepath, **kwargs)
    
    def export_csv(self, filepath, **kwargs):
        """Export to CSV"""
        self.df.to_csv(filepath, index=False, **kwargs)
        print(f"✓ CSV exported: {filepath}")
        return filepath
    
    def export_excel(self, filepath, sheet_name='Data', styled=False, **kwargs):
        """Export to Excel"""
        if styled:
            # Export with styling
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            
            # Add headers with style
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, col_name in enumerate(self.df.columns, 1):
                cell = sheet.cell(row=1, column=col_num, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
            
            # Add data
            for row_num, row_data in enumerate(self.df.values, 2):
                for col_num, value in enumerate(row_data, 1):
                    sheet.cell(row=row_num, column=col_num, value=value)
            
            workbook.save(filepath)
        else:
            self.df.to_excel(filepath, sheet_name=sheet_name, index=False, **kwargs)
        
        print(f"✓ Excel exported: {filepath}")
        return filepath
    
    def export_json(self, filepath, orient='records', include_metadata=True, **kwargs):
        """Export to JSON"""
        if include_metadata:
            data = {
                'metadata': {
                    'export_date': datetime.now().isoformat(),
                    'record_count': len(self.df),
                    'columns': list(self.df.columns)
                },
                'data': self.df.to_dict(orient=orient)
            }
            with open(filepath, 'w') as f:
                json.dump(data, f, indent=2, default=str)
        else:
            self.df.to_json(filepath, orient=orient, indent=2, date_format='iso', **kwargs)
        
        print(f"✓ JSON exported: {filepath}")
        return filepath
    
    def export_all(self, base_filename):
        """Export to all formats at once"""
        print(f"\n📦 Exporting '{base_filename}' to all formats...")
        self.export(base_filename, format='all')
        print(f"✅ All exports completed!")

# Example usage of DataExporter class
print("\n" + "="*70)
print("USING DataExporter CLASS")
print("="*70)

exporter = DataExporter(df, output_dir='exports')

# Export to all formats
exporter.export_all('customer_data')

# Or export individually
# exporter.export('my_data', format='csv')
# exporter.export('my_data', format='excel', styled=True)
# exporter.export('my_data', format='json', include_metadata=True)

# ============================================================================
# SUMMARY
# ============================================================================
print("\n" + "="*70)
print("EXPORT SUMMARY")
print("="*70)
print("""
📁 All files exported to 'exports/' directory

CSV Formats:
  ✓ Basic CSV
  ✓ Custom separator (semicolon)
  ✓ Selected columns only
  ✓ UTF-8 with BOM encoding
  ✓ Filtered data
  ✓ Custom date format
  ✓ Append mode

Excel Formats:
  ✓ Basic Excel
  ✓ Multiple sheets
  ✓ Formatted (colors, borders)
  ✓ Conditional formatting
  ✓ With formulas

JSON Formats:
  ✓ Records format (array)
  ✓ Index format
  ✓ Columns format
  ✓ Custom nested structure
  ✓ Grouped data
  ✓ JSON Lines (JSONL)

💡 Use the DataExporter class for production code!
""")

print("\n✅ All export examples completed successfully!")
print(f"📂 Check the 'exports/' directory for all generated files")